import pandas as pd
from matplotlib import pyplot as plt
import openpyxl
from tdigest import TDigest
import json

# open Dat sets for T-Digest
df = openpyxl.load_workbook('Dat sets for T-Digest (1).xlsx')
data = df.active

# Chose the column to read
print(f"Choose the column to read,\n1 for \"{data.cell(row=5,column = 2).value}\",\n2 for \"{data.cell(row=5,column = 3).value}\",\n3 for \"{data.cell(row=5,column = 4).value}\",\n4 for \"{data.cell(row=5,column = 5).value}\",\n5 for \"{data.cell(row=5,column = 6).value}\",\nand 6 for \"{data.cell(row=5,column = 7).value}\"")
response = int(input())
if response > 6 or response < 1:
    print("Invalid input")
    exit(1)
col = 1 + response

digest = TDigest()
# array = []
for i in range(14, 14+data.cell(row=1, column=col).value):
    # array.append(data.cell(column=col, row=i).value)
    digest.update(data.cell(column=col, row=i).value)
digest.compress()

with open("output.json", "w+") as outfile:
    json.dump(digest.to_dict(), outfile, indent=4)

print("Done, centroids are saved in output.json")
